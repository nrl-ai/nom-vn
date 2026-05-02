"""Format-preserving ``.xlsx`` translation walker.

Walks every sheet's cells via ``openpyxl``. Cells with a string value
are translated in place; numbers, dates, formulas (``cell.data_type ==
'f'``), and empty cells pass through untouched. Cell formatting (font,
fill, border, number-format), merged-cell ranges, named ranges, and
sheet structure are preserved.

Trade-offs:

- **Sheet names** are not translated in v0 — formula references like
  ``=Sheet1!A1`` would break. Revisit if user demand surfaces; the fix
  is to walk every formula and rewrite sheet refs after rename.
- **Comments** on cells (``cell.comment.text``) are translated.
- **Data validation lists / dropdowns** are NOT translated in v0 —
  they live in a separate XML stream openpyxl exposes as
  ``sheet.data_validations`` and need a per-rule walker.
- **Hyperlinks**: the *display text* of a hyperlinked cell is the cell
  value, which IS translated. The URL itself is left alone.
"""

from __future__ import annotations

import contextlib
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING, Any

from nom.translate.base import Translator

if TYPE_CHECKING:
    pass

__all__ = ["XlsxTranslationStats", "translate_xlsx"]


@dataclass(frozen=True, slots=True)
class XlsxTranslationStats:
    """Summary of a single :func:`translate_xlsx` run."""

    cells_translated: int
    cells_skipped: int
    cells_failed: int
    chars_in: int
    chars_out: int


def translate_xlsx(
    src: Path | str,
    dst: Path | str,
    translator: Translator,
) -> XlsxTranslationStats:
    """Translate string cells in a ``.xlsx``. Source untouched; ``dst``
    written fresh."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "Translating .xlsx requires openpyxl. Install with: pip install nom-vn[doc]"
        ) from exc

    src_path = Path(src)
    dst_path = Path(dst)
    if not src_path.exists():
        raise FileNotFoundError(f"xlsx source not found: {src_path}")

    # data_only=False keeps formulas intact; read_only=False so we can
    # mutate cell values and save.
    wb = load_workbook(str(src_path))
    counts = _Counts()

    for sheet in wb.worksheets:
        _translate_sheet(sheet, translator, counts)

    dst_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dst_path))
    return counts.freeze()


@dataclass
class _Counts:
    translated: int = 0
    skipped: int = 0
    failed: int = 0
    chars_in: int = 0
    chars_out: int = 0

    def freeze(self) -> XlsxTranslationStats:
        return XlsxTranslationStats(
            cells_translated=self.translated,
            cells_skipped=self.skipped,
            cells_failed=self.failed,
            chars_in=self.chars_in,
            chars_out=self.chars_out,
        )


def _translate_sheet(sheet: Any, translator: Translator, counts: _Counts) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            value = cell.value
            if value is None or not isinstance(value, str):
                continue
            # Skip formula cells — `cell.data_type == 'f'` covers them
            # but openpyxl also exposes a leading '=' on the value when
            # the workbook was loaded with formulas. Belt + suspenders.
            if cell.data_type == "f" or value.startswith("="):
                continue
            stripped = value.strip()
            if not stripped:
                counts.skipped += 1
                continue
            counts.chars_in += len(value)
            try:
                translated = translator.translate(value)
            except Exception:
                counts.failed += 1
                continue
            cell.value = translated
            counts.chars_out += len(translated)
            counts.translated += 1

            # Cell comment, if any
            if cell.comment is not None and cell.comment.text:
                with contextlib.suppress(Exception):
                    cell.comment.text = translator.translate(cell.comment.text)
